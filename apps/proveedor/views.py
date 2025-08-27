from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
import json

# Para la generación de Excel
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# Importación de modelos y formularios de tu app
from apps.proveedor.models import Proveedor, Prenda
from apps.proveedor.forms import ProveedorForm, PrendaForm


# Vistas basadas en funciones
@login_required
def ProveedorHome(request):
    return render(request, "proveedor/proveedor_home.html")

def custom_logout(request):
    logout(request)
    return redirect('login') 

@login_required
def proveedores_download(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    columnas = [
        "Nit", "Nombre", "Teléfono", "Dirección", "Ciudad", "Contacto", "Correo",
        "Móvil", "Plazo", "Descuento", "Contado", "Categorías"
    ]
    ws.append(columnas)

    proveedores = Proveedor.objects.all()
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for col_idx, column_name in enumerate(columnas, start=1):
        ws.cell(row=1, column=col_idx).fill = header_fill

    for p in proveedores:
        categorias = p.Categorias or "" 
        fila = [
            p.Nit,
            p.Nombre,
            p.Telefono,
            p.Direccion,
            p.Ciudad,
            p.Contacto,
            p.Correo,
            p.Movil,
            p.Plazo,
            p.Descuento or "",
            p.get_Contado_display(),
            categorias
        ]
        ws.append(fila)

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response


# Vistas basadas en clases para Proveedor
class ProveedorListView(LoginRequiredMixin, ListView):
    model = Proveedor
    template_name = 'proveedor/Proveedor.html'
    context_object_name = 'proveedores'
    login_url = 'login'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset()
        ciudad = self.request.GET.get('ciudad')
        contacto = self.request.GET.get('contacto')
        nombre = self.request.GET.get('nombre')
        categoria = self.request.GET.get('categoria')
        orden = self.request.GET.get('orden')

        if ciudad:
            queryset = queryset.filter(Ciudad__icontains=ciudad)
        if contacto:
            queryset = queryset.filter(Contacto__icontains=contacto)
        if nombre:
            queryset = queryset.filter(Nombre__icontains=nombre)
        if categoria:
            queryset = queryset.filter(Categorias__icontains=categoria)
        if orden == 'asc':
            queryset = queryset.order_by('ID')
        elif orden == 'desc':
            queryset = queryset.order_by('-ID')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = self.request.GET
        context['ciudades'] = Proveedor.objects.values_list('Ciudad', flat=True).distinct()
        context['contactos'] = Proveedor.objects.values_list('Contacto', flat=True).distinct()
        context['nombres'] = Proveedor.objects.values_list('Nombre', flat=True).distinct()

        categorias_unicas = set()
        for proveedor in Proveedor.objects.exclude(Categorias__isnull=True).exclude(Categorias=''):
            categorias_unicas.update([cat.strip() for cat in proveedor.get_categorias_list()])

        context['categorias'] = sorted(categorias_unicas)
        return context

class ProveedorCreateView(LoginRequiredMixin, CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor/Proveedor_crear.html'
    context_object_name = 'Proveedor'
    login_url = "login"
    success_url = reverse_lazy('proveedores_list') 

class ProveedorUpdateView(LoginRequiredMixin, UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'proveedor/Proveedor_editar.html'
    context_object_name = 'Proveedor'
    login_url = "login"
    success_url = reverse_lazy('proveedores_list') 
    pk_url_kwarg = 'ProveedorID'

class ProveedorDeleteView(LoginRequiredMixin, DeleteView):
    model = Proveedor
    template_name = 'proveedor/proveedor_eliminar.html'
    success_url = reverse_lazy('proveedores_list')
    pk_url_kwarg = 'ProveedorID'

class ProveedorDetailView(LoginRequiredMixin, DetailView):
    model = Proveedor
    template_name = 'proveedor/proveedor_detalle.html'
    context_object_name = 'Proveedor'
    pk_url_kwarg = 'ProveedorID'






# Vistas basadas en clases para Prenda
class PrendaCreateView(LoginRequiredMixin, CreateView):
    model = Prenda
    form_class = PrendaForm
    template_name = 'proveedor/prenda_crear.html'
    login_url = 'login'
    success_url = reverse_lazy('proveedores_list')


class PrendaUpdateView(LoginRequiredMixin, UpdateView):
    model = Prenda
    form_class = PrendaForm
    template_name = 'proveedor/prenda_editar.html'
    context_object_name = 'prenda'
    login_url = "login"
    success_url = reverse_lazy('proveedores_list') 

class PrendaDeleteView(LoginRequiredMixin, DeleteView):
    model = Prenda
    template_name = 'proveedor/prenda_eliminar.html'
    context_object_name = 'prenda'
    success_url = reverse_lazy('proveedores_list')
    login_url = "login"