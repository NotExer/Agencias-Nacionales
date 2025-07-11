from django.shortcuts import render
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from apps.comercial.models import Clientes
from apps.comercial.forms import Clientesform
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.http import HttpResponseRedirect
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

@login_required
def ComercialHome(request):
    return render(request, "comercial/comercial_home.html")

def custom_logout(request):
    logout(request)
    return redirect('login') 

class ClientesListView(LoginRequiredMixin, ListView):
    model = Clientes
    template_name = 'comercial/clientes.html'
    context_object_name = 'clientes'
    login_url = 'login'
    paginate_by = 25

    def get_queryset(self):
        queryset = super().get_queryset()

        ciudad = self.request.GET.get('ciudad')
        estado = self.request.GET.get('estado')
        vendedor = self.request.GET.get('vendedor')
        razon = self.request.GET.get('razon')
        orden_fecha = self.request.GET.get('orden')

        if ciudad:
            queryset = queryset.filter(Ciudad__icontains=ciudad)

        if estado:
            queryset = queryset.filter(Estado__icontains=estado)

        if vendedor:
            queryset = queryset.filter(Vendedor__icontains=vendedor)

        if razon:
            queryset = queryset.filter(Razon__icontains=razon)

        if orden_fecha == 'asc':
            queryset = queryset.order_by('FechaUltimoPedido')
        elif orden_fecha == 'desc':
            queryset = queryset.order_by('-FechaUltimoPedido')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtros'] = self.request.GET  # Para conservar valores en el formulario
        context['ciudades'] = Clientes.objects.values_list('Ciudad', flat=True).distinct()
        context['estados'] = Clientes.objects.values_list('Estado', flat=True).distinct()
        context['vendedores'] = Clientes.objects.values_list('Vendedor', flat=True).distinct()
        return context


class ClientesCreateView(LoginRequiredMixin, CreateView):
    model = Clientes
    form_class = Clientesform
    template_name = 'comercial/clientes_crear.html'
    context_object_name = 'clientes'
    login_url = "login"
    success_url = reverse_lazy('clientes_home') 


class ClientesUpdateView(LoginRequiredMixin, UpdateView):
    model = Clientes
    form_class = Clientesform
    template_name = 'comercial/clientes_editar.html'
    context_object_name = 'clientes'
    login_url = "login"
    success_url = reverse_lazy('clientes_home') 
    pk_url_kwarg = 'ClienteID'



class ClientesDeleteView(LoginRequiredMixin, DeleteView):
    model = Clientes
    template_name = 'comercial/clientes_eliminar.html'
    success_url = reverse_lazy('clientes_home')
    pk_url_kwarg = 'ClienteID'



class ClienteDetailView(LoginRequiredMixin, DetailView):
    model = Clientes
    template_name = 'comercial/clientes_detalle.html'
    context_object_name = 'clientes'
    pk_url_kwarg = 'ClienteID'

    
@login_required
def clientes_download(request):
    # Crear un libro y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados
    columnas = [
        "Nit", "Razón Social", "Nombre Comercial", "Sigla", "Ciudad",
        "Teléfono", "Celular", "Contacto", "Cargo", "Correo",
        "Sector", "Página", "Estado", "Vendedor",
        "Comentario"
    ]
    ws.append(columnas)

    # Obtener todos los clientes
    clientes = Clientes.objects.all()

    # Rellenar datos
    for idx, c in enumerate(clientes, start=1):
        fila = [
            c.Nit, c.Razon, c.Nombre, c.Sigla, c.Ciudad,
            c.Telefono, c.Celular, c.Contacto, c.Cargo, c.Correo,
            c.Sector, c.Pagina, c.Estado, c.Vendedor,
            c.Comentario
        ]
        ws.append(fila)

        # Colorear la primera fila de contenido (idx == 1 porque start=1)
        if idx == 1:
            for col in range(1, len(fila) + 1):
                ws.cell(row=1, column=col).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Ajustar ancho de columnas automáticamente
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col_cells:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = max_length + 2  # Padding extra
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Preparar respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response

